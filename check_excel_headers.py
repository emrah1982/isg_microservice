import openpyxl

file_path = r"C:\Users\User\Desktop\isg_microservice\ReactApp\public\template\AKDO PERSONEL VERİ TAKİP TABLOSU 15.12.2025 İSG.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print("Excel Başlıkları (İlk Satır):")
    print("-" * 80)
    
    for idx, cell in enumerate(ws[1], 1):
        value = cell.value
        if value:
            # Normalize edilmiş hali
            normalized = value.strip().lower()
            normalized = normalized.replace("ı", "i").replace("ş", "s").replace("ç", "c").replace("ö", "o").replace("ü", "u").replace("ğ", "g")
            normalized = ''.join(c for c in normalized if c.isalnum())
            
            print(f"Sütun {idx}: '{value}'")
            print(f"  Normalize: '{normalized}'")
            print()
    
    print("\nİlk 3 Satır Verisi:")
    print("-" * 80)
    for row_idx in range(2, min(5, ws.max_row + 1)):
        row = ws[row_idx]
        print(f"Satır {row_idx}:")
        for idx, cell in enumerate(row[:10], 1):  # İlk 10 sütun
            if cell.value:
                print(f"  Sütun {idx}: {cell.value}")
        print()
        
except FileNotFoundError:
    print(f"Dosya bulunamadı: {file_path}")
except Exception as e:
    print(f"Hata: {e}")
    import traceback
    traceback.print_exc()
